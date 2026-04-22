import sys, os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side

import sys, os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def xlsxToJson(file_name, sheetNumber):
    """
    Lee un archivo Excel y lo convierte a una lista de diccionarios (JSON).
    Usa data_only=True para obtener los valores calculados y aplica redondeo
    para evitar artefactos de precisión de punto flotante de Excel.
    """
    # data_only=True es crucial para leer el valor resultante de una fórmula en lugar de la fórmula misma. Esto asegura que obtenemos el valor que se muestra en Excel
    wb = load_workbook(filename=file_name, data_only=True)
    sheet = wb[wb.sheetnames[sheetNumber-1]]
    
    headers = [cell.value for cell in sheet[1]]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for idx, value in enumerate(row):
            header = headers[idx]
            
            # Limpieza de precisión binaria (ej: 0.31273199999999995 -> 0.312732)
            # Redondear a 10 decimales suele ser suficiente para corregir el ruido de punto flotante
            # sin alterar datos de alta precisión.
            if isinstance(value, (float)):
                value = round(value, 10)
            
            row_dict[header] = value
        data.append(row_dict)
        
    return data

def generateInstanceOfExcelWithJson(data):
    """
    Genera un archivo Excel formateado a partir de una lista de diccionarios.
    Optimiza el ajuste de columnas y el uso de memoria.
    """
    if not data:
        return

    headers = list(data[0].keys())
    wb = Workbook()
    ws = wb.active
    
    # Estilos predefinidos
    header_fill = PatternFill(start_color='74ac44', end_color='74ac44', fill_type="solid")
    row_fill_alt = PatternFill(start_color='e2efda', end_color='e2efda', fill_type="solid")
    common_alignment = Alignment(horizontal='center', vertical='center')
    common_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color='ffffff')

    # Escritura de encabezados
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.alignment = common_alignment
        cell.border = common_border
        cell.font = header_font

    # Escritura de datos
    for row_idx, row_data in enumerate(data, 2):
        is_alternate = row_idx % 2 != 0
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header)
            # Manejo de valores nulos o "0" según lógica original
            display_val = '' if str(val) == "0" else val
            
            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            cell.alignment = common_alignment
            cell.border = common_border
            if is_alternate:
                cell.fill = row_fill_alt

    # Auto-ajuste de columnas optimizado (fuera del bucle de filas)
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                str_val = str(cell.value) if cell.value is not None else ""
                if len(str_val) > max_length:
                    max_length = len(str_val)
            except:
                pass
        # Factor de ajuste para el ancho de columna
        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    ws.freeze_panes = "B2"
    
    # Ruta temporal y ejecución
    temp_dir = os.environ.get('TEMP', os.environ.get('USERPROFILE'))
    temp_route = os.path.join(temp_dir, 'temp_output.xlsx')
    
    try:
        wb.save(temp_route)
        os.system(f"start EXCEL.EXE {temp_route}")
    except PermissionError:
        print(f"Error: No se pudo guardar. Cierre el archivo {temp_route} antes de reintentar.")

#generateInstanceOfExcelWithJson(xlsxToJson('temp.xlsx',1))
